import discord
import os
from discord.ext import commands
import Randomize
import Crawling
import Search
import openpyxl
import translation

client = commands.Bot(command_prefix = '.')

@client.event
async def on_ready():
    print('Bot is ready.')

@client.event
async def on_message(message):

    channel = message.channel
    if message.content.startswith('.폭스야'):
        await client.send_message(channel, '왜 :question:')

    if message.content.startswith('.안녕'):
        await client.send_message(channel, '안녕~ :sunglasses:')

    if message.content.startswith('.Champion'):
        await client.send_message(channel, '가렌 갈리오 갱플 그라가스 그브 나르 나미 나서스 노틸 녹턴 누누 니달리 니코 '
                                           '다리 다이애나 드븐 라이즈 라칸 람머 럭스 럼블 레넥 레오나 렉사이 렝가 루시안 '
                                           '룰루 르블랑 리신 리븐 리산 마이 마오 말자 말파 모데 몰가 문도 미포 바드 '
                                           '바루스 바이 베이가 베인 벨코즈 볼베 브랜드 브라움 블라디 블츠 빅토르 뽀삐 '
                                           '사이온 사일러스 샤코 세주 소나 소라카 쉔 쉬바나 스웨인 스카너 시비르 짜오'
                                           '신드라 신지드 쓰레쉬 아리 아무무 아우솔 아이번 아지르 아칼리 아트 알리 애니'
                                           '애니비아 애쉬 야스오 에코 엘리스 오공 오른 오리 올라프 요릭 우디르 우르곳'
                                           '워윅 이렐리아 이블린 이즈 일라 자르반 자야 자이라 자크 잔나 잭스 제드 제라스'
                                           '제이스 조이 직스 진 질리언 징크스 초가스 카르마 카밀 카사딘 카서스 카시'
                                           '카이사 카직스 칼리 카타 케넨 케틀 케인 케일 코그모 코르키 퀸 클레드 킨드'
                                           '타릭 탈론 탈리야 켄치 트런들 트타 트린 트페 트위치 티모 파이크 판테 피들'
                                           ' 피오라 피즈 딩거 헤카림')

    if message.content.startswith('.에코'):
         msg = message.content.split()
         output = ''
         for word in msg[1:]:
                output += word
                output += ' '
         await client.send_message(channel, output)

    if message.content.startswith(".명령어"):
        channel = message.channel
        embed = discord.Embed(
        title = 'Fox 봇 - 명령어',
        description = 'FOX 봇은 클장 자체 제작으로 원하시는 사항이 있으면 말씀해주세요',
        colour = discord.Colour.blue()
        )
        embed.add_field(name='1 ', value='.이모티콘 - 랜덤 이모티콘~', inline=False)
        embed.add_field(name='2', value='.주사위 - 랜덤 숫자', inline=False)
        embed.add_field(name='3', value='.영화 - 최신 영화 순위', inline=False)
        embed.add_field(name='4', value='.실검 - 네이버 실검 10위', inline=False)
        embed.add_field(name='5', value='.이미지 - 네이버 이미지 검색', inline=False)
        embed.add_field(name='6', value='.전적 롤아이디 - 롤 전적정보', inline=False)
        embed.add_field(name='7', value='.모스트 롤아이디 - 롤 모스트정보', inline=False)
        embed.add_field(name='8', value='.럼블 - OPGG기준 [탑, 정글, 미드]에서 럼블 순위 검색', inline=False)
        embed.add_field(name='9', value='.탑 - OPGG기준 탑 op챔프 5위', inline=False)
        embed.add_field(name='10', value='.정글 - OPGG기준 정글 op챔프 5위', inline=False)
        embed.add_field(name='11', value='.미드 - OPGG기준 미드 op챔프 5위', inline=False)
        embed.add_field(name='12', value='.원딜 - OPGG기준 원딜 op챔프 5위', inline=False)
        embed.add_field(name='13', value='.서폿 - OPGG기준 서폿 op챔프 5위', inline=False)
        await client.send_message(channel, embed=embed)

    ## Read & Write FILE
    if message.content.startswith(".쓰기"):
        msg = message.content.split("/")
        print(msg[1])
        file = open(msg[1] + ".txt", "w")
        file.write(msg[2])
        file.close()

    if message.content.startswith(".읽기"):
        msg = message.content.split(" ")
        file = open(msg[1] + ".txt")
        await client.send_message(message.channel, file.read())
        file.close()

    if message.content.startswith(".등록"):
        saveMessage = message.content.split("/")
        file = openpyxl.load_workbook("유저정보.xlsx")
        sheet = file.active
        for i in range(1, 101):
            if sheet["A" + str(i)].value == "-" or sheet["A" + str(i)].value == saveMessage[1]:
                print(saveMessage[1])
                sheet["A" + str(i)].value = saveMessage[1]
                sheet["B" + str(i)].value = saveMessage[2]
                sheet["C" + str(i)].value = saveMessage[3]
                sheet["D" + str(i)].value = saveMessage[4]
                sheet["E" + str(i)].value = saveMessage[5]
                break
        file.save("유저정보.xlsx")
        await  client.send_message(message.channel, "저장완료!")

    if message.content.startswith(".삭제"):
        saveMessage = message.content.split("/")
        file = openpyxl.load_workbook("유저정보.xlsx")
        sheet = file.active
        for i in range(1, 101):
            if sheet["A" + str(i)].value == saveMessage[1]:
                print(saveMessage[1])
                sheet["A" + str(i)].value = "-"
                sheet["B" + str(i)].value = "-"
                sheet["C" + str(i)].value = "-"
                sheet["D" + str(i)].value = "-"
                sheet["E" + str(i)].value = "-"
                break
        file.save("유저정보.xlsx")
        await  client.send_message(message.channel, "삭제완료!")

    if message.content.startswith(".검색"):
        searchMessage = message.content.split(" ")
        file = openpyxl.load_workbook("유저정보.xlsx")
        sheet = file.active
        SearchCheck = 0;
        info = []
        for i in range(1, 101):
            if sheet["A" + str(i)].value == searchMessage[1]:
                info.append(sheet["A" + str(i)].value)
                info.append(sheet["B" + str(i)].value)
                info.append(sheet["C" + str(i)].value)
                info.append(sheet["D" + str(i)].value)
                info.append(sheet["E" + str(i)].value)
                SearchCheck = 1;
                break
        if SearchCheck == 1:
            embed = discord.Embed(
                title='유저 정보',
                description= info[0] + '의 등록정보',
                colour=discord.Color.purple()
            )
            embed.add_field(name='Team', value=info[1], inline=False)
            embed.add_field(name='Tier', value=info[2], inline=False)
            embed.add_field(name='Line', value=info[3], inline=False)
            embed.add_field(name='Champ', value=info[4], inline=False)
            await client.send_message(message.channel, embed = embed)
        else:
            await client.send_message(message.channel,  '검색한 유저 정보가 없습니다.')

    if message.content.startswith(".팀"):
        searchTeam = message.content.split(" ")
        file = openpyxl.load_workbook("유저정보.xlsx")
        sheet = file.active

        UserName = []
        UserTier = []
        UserLine = []
        UserChamp = []

        if searchTeam[1] == "A" or "B" or "C" or "D":
            for i in range(1, 101):
                if sheet["B" + str(i)].value == searchTeam[1]:
                    UserName.append(sheet["A" + str(i)].value)
                    UserTier.append(sheet["C" + str(i)].value)
                    UserLine.append(sheet["D" + str(i)].value)
                    UserChamp.append(sheet["E" + str(i)].value)

            embed = discord.Embed(
                title= searchTeam[1] + "팀 정보",
                description= '선수 등록정보',
                colour=discord.Color.purple()
            )
            print(str(len(UserName)))

            for i in range(0, len(UserName)):
                embed.add_field(name=UserName[i], value="-      " + UserLine[i] + " " + UserTier[i] + " " + UserChamp[i], inline=False)

            await client.send_message(message.channel, embed=embed)

        else:
            await client.send_message(message.channel, '검색한 팀 정보가 없습니다.')


    if message.content.startswith(".팀나누기"):
        Totalcontent = message.content[6:]
        splitcontent = Totalcontent.split("/")
        Users = splitcontent[0]
        Teams = splitcontent[1]
        User = Users.split(" ")
        Team = Teams.split(" ")
        Randomize.Get_Shuffle(Team)
        for i in range(0, len(User)):
            await client.send_message(message.channel, User[i] + "-->" + Team[i])

    ## Search for 'League of Legend'
    if message.content.startswith(".전적"):
        learn = message.content.split(" ")
        ##location = learn[1]
        channel = message.channel
        summonerName = ''
        for wr in learn[1:]:
            summonerName += wr

        location = summonerName

        embed = discord.Embed(
            title='전적 정보',
            description=summonerName + '의 솔로랭크 전적입니다.',
            colour=discord.Colour.green()
        )

        Record_embed = Crawling.Get_Record(location, embed)
        await client.send_message(channel, embed=Record_embed)

    if message.content.startswith('.럼블'):
        channel = message.channel

        embed = discord.Embed(
            title='RUMBEL',
            description='럼블의 현실',
            colour=discord.Colour.red()
        )

        Rumble_embed = Crawling.Get_Rumble(embed)
        await client.send_message(channel, embed=Rumble_embed)

    if message.content.startswith(".챔프"):
        learn = message.content.split(" ")
        channel = message.channel
        champName = learn[1]

        EngN = translation.Trans_ChampN(champName)



        if EngN  == "":
            await client.send_message(channel, '챔프 이름을 확인해주세요.')
        else :
            embed = discord.Embed(
                    title='Champion Info',
                    description= champName + '의 순위 정보입니다.',
                    colour=discord.Color.dark_orange()
            )

            Line_num = translation.Line_ChampN(EngN)

            Champ_embed = Crawling.Get_Champion(EngN, Line_num, embed)
            await client.send_message(channel, embed=Champ_embed)

    if message.content.startswith(".모스트"):
        learn = message.content.split(" ")
        channel = message.channel
        summonerName = ''
        for wr in learn[1:]:
            summonerName += wr

        location = summonerName

        embed = discord.Embed(
            title='모스트 정보',
            description=summonerName + '의 롤 모스트입니다.',
            colour=discord.Colour.gold()
        )

        Most_embed = Crawling.Get_Most(location, embed)
        await client.send_message(channel, embed=Most_embed)

    if message.content.startswith('.주사위'):
        Dice = Randomize.Get_Dice()
        await client.send_message(message.channel, embed=discord.Embed(description=':game_die: ' + Dice))

    if message.content.startswith('.이모티콘'):
        Emoji = Randomize.Get_emoji()
        await client.send_message(message.channel, embed=discord.Embed(description=Emoji))
        # 랜덤 이모티콘을 메시지로 출력합니다.

    if message.content.startswith('.영화'):
        embed = discord.Embed(
            title = "영화순위",
            description = "영화순위입니다.",
            colour = discord.Color.red()
            )
        Ranking_embed = Crawling.Get_MRanking(embed)
        await client.send_message(message.channel, embed=Ranking_embed)

    if message.content.startswith(".이미지"):
        img = message.content.split(" ")

        SearchInfo = ''
        for wr in img[1:]:
            SearchInfo += wr

        imgsrc = Search.get_image(SearchInfo)
        await client.send_message(message.channel, imgsrc)

    if message.content.startswith('.실검'):
        embed = discord.Embed(
            title='네이버 실시간 검색어',
            description='실시간검색어',
            colour=discord.Colour.green()
        )
        Crawling.Get_NaverChart(embed)
        await client.send_message(message.channel, embed=embed)

    if message.content.startswith('.탑'):
        embed = discord.Embed(
            title = "탑 TOP",
            description = "탑 챔프 티어순위",
            colour = discord.Color.red()
            )
        Tier_embed = Crawling.Get_topTier(embed)
        await client.send_message(message.channel, embed=Tier_embed)

    if message.content.startswith('.정글'):
        embed = discord.Embed(
            title="정글 JG",
            description="정글 챔프 티어순위",
            colour=discord.Color.orange()
        )
        Tier_embed = Crawling.Get_jgTier(embed)
        await client.send_message(message.channel, embed=Tier_embed)

    if message.content.startswith('.미드'):
        embed = discord.Embed(
            title="미드 MID",
            description="미드 챔프 티어순위",
            colour=discord.Color.purple()
        )
        Tier_embed = Crawling.Get_midTier(embed)
        await client.send_message(message.channel, embed=Tier_embed)

    if message.content.startswith('.원딜'):
        embed = discord.Embed(
            title="원딜 AD",
            description="원딜 챔프 티어순위",
            colour=discord.Color.green()
        )
        Tier_embed = Crawling.Get_adcTier(embed)
        await client.send_message(message.channel, embed=Tier_embed)

    if message.content.startswith('.서폿'):
        embed = discord.Embed(
            title="서폿 SUP",
            description="서폿 챔프 티어순위",
            colour=discord.Color.blue()
        )
        Tier_embed = Crawling.Get_supTier(embed)
        await client.send_message(message.channel, embed=Tier_embed)

access_token = os.environ["BOT_TOKEN"]
client.run(access_token)
